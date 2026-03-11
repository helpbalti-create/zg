"""
Management command для импорта бюджета из Excel-файла UMCOR/CWS.

УСТАНОВКА (один раз):
  budget/management/__init__.py          <- пустой файл
  budget/management/commands/__init__.py <- пустой файл
  budget/management/commands/import_budget_excel.py <- этот файл

ИСПОЛЬЗОВАНИЕ:
  python manage.py import_budget_excel "путь/к/файлу.xlsx" --user email@example.com
  python manage.py import_budget_excel "путь/к/файлу.xlsx" --user email@example.com --dry-run
  python manage.py import_budget_excel "путь/к/файлу.xlsx" --user email@example.com --force
"""

import re
from datetime import date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    raise ImportError("Установи: pip install openpyxl")


SECTION_PATTERN = re.compile(
    r'SECTION\s+([A-Z](?:\.\d+)?)\s*[:\-–]?\s*(.*)', re.IGNORECASE
)
SUB_TOTAL_PATTERN = re.compile(
    r'sub.?total|total\s+direct|total\s+program|total\s+grant|total\s+other',
    re.IGNORECASE
)
MONTH_RU = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4,
    'май': 5, 'июнь': 6, 'июль': 7, 'август': 8,
    'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12,
}
SKIP_HEADERS = {
    'position', 'support personnel', 'equipment',
    'office rental and supplies', 'travel', 'domestic', 'international',
    'communication', 'other operational costs',
    'visibility & publication costs', 'monitoring & evaluation costs',
    'in-kind donation/matching funds',
}


def to_decimal(val, default=Decimal('0')):
    if val is None:
        return default
    try:
        return Decimal(str(val)).quantize(Decimal('0.01'))
    except InvalidOperation:
        return default


def parse_date_ru(text):
    text = str(text).strip()
    m = re.search(r'(\d+)\s+([а-яёА-ЯЁ]+)\s+(\d{4})', text)
    if m:
        month = MONTH_RU.get(m.group(2).lower())
        if month:
            return date(int(m.group(3)), month, int(m.group(1)))
    m2 = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', text)
    if m2:
        return date(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    return None


def find_detail_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        text = ' '.join(
            str(c.value) for row in ws.iter_rows(max_row=140)
            for c in row if c.value
        )
        if 'SECTION B' in text.upper() and 'BUDGET DESCRIPTION' in text.upper():
            return ws, name
    raise CommandError(
        f"Не найден лист с детальным бюджетом. Листы: {wb.sheetnames}"
    )


def parse_excel(filepath, sheet_name=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Лист '{sheet_name}' не найден. Доступные: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws, sheet_name = find_detail_sheet(wb)

    rows = list(ws.iter_rows(values_only=True))

    # Мета-данные
    project_name = 'CWS 2025 Ukraine Response Program – Serbia and Moldova'
    donor = 'UMCOR / Church World Service'
    project_code = 'UMCOR-5'
    start_date = date(2026, 2, 1)
    end_date = date(2026, 8, 1)
    total_budget = Decimal('58999')

    for row in rows[:20]:
        cells = [str(c).strip() if c else '' for c in row]
        full = ' '.join(cells)

        if 'Project Title:' in full:
            m = re.search(r'Project Title:\s*(.+)', full)
            if m:
                project_name = m.group(1).strip()

        for cell in cells:
            if 'Дата начала' in cell:
                d = parse_date_ru(cell.replace('Дата начала:', '').strip())
                if d:
                    start_date = d
            if 'Дата окончания' in cell:
                d = parse_date_ru(cell.replace('Дата окончания:', '').strip())
                if d:
                    end_date = d

        if 'TOTAL INCOME' in full:
            for c in row:
                v = to_decimal(c, None)
                if v and v > 100:
                    total_budget = v
                    break

    # Парсинг секций и статей
    # Колонки листа: 0=№  1=CODE  2=DESCRIPTION  3=UNIT  4=QTY  5=UNIT_COST  6=FREQ  7=TOTAL  8=NOTES
    sections = []
    current_section = None
    category_order = 0

    for row in rows[18:]:
        col_b = str(row[1]).strip() if row[1] is not None else ''
        col_c = str(row[2]).strip() if row[2] is not None else ''
        col_d = str(row[3]).strip() if row[3] is not None else ''
        full_line = col_b + ' ' + col_c

        # Новая секция
        sec_m = SECTION_PATTERN.match(col_b)
        if sec_m:
            sec_code = sec_m.group(1).upper()
            sec_name = sec_m.group(2).strip() or col_c.strip()
            if '#REF' in sec_name:
                sec_name = col_c.strip()
            current_section = {
                'code': sec_code,
                'name': sec_name or f'Section {sec_code}',
                'categories': [],
            }
            sections.append(current_section)
            category_order = 0
            continue

        # Пропускаем служебные строки
        if SUB_TOTAL_PATTERN.search(full_line):
            continue
        if '#REF' in full_line:
            continue
        if col_c.strip().rstrip('.').lower() in SKIP_HEADERS:
            continue

        # Строка статьи: col[0] = числовой порядковый номер
        row_num = row[0]
        if not isinstance(row_num, (int, float)):
            continue

        code = col_b if col_b not in ('None', '') else ''
        name = col_c
        unit = col_d if col_d not in ('None', '') else 'Lump sum'
        qty = to_decimal(row[4], Decimal('1'))
        unit_cost = to_decimal(row[5], Decimal('0'))
        freq = to_decimal(row[6], Decimal('1'))
        total = to_decimal(row[7], Decimal('0'))
        notes = str(row[8]).strip() if row[8] and str(row[8]) not in ('#REF!', 'None') else ''

        if not name or name == 'None' or '#REF' in name:
            continue
        if total == Decimal('0') and unit_cost == Decimal('0'):
            continue

        if current_section is None:
            current_section = {'code': 'B', 'name': 'DIRECT PROGRAM COSTS', 'categories': []}
            sections.append(current_section)

        category_order += 1
        current_section['categories'].append({
            'code': code,
            'name': name,
            'unit_measure': unit,
            'quantity': qty,
            'unit_cost': unit_cost,
            'frequency': freq,
            'allocated_amount': total if total > 0 else qty * unit_cost * freq,
            'notes': notes,
            'order': category_order,
        })

    return {
        'name': project_name,
        'donor': donor,
        'project_code': project_code,
        'start_date': start_date,
        'end_date': end_date,
        'total_budget': total_budget,
        'sections': sections,
    }


class Command(BaseCommand):
    help = 'Импорт бюджетного проекта из Excel-файла UMCOR/CWS'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Путь к .xlsx файлу')
        parser.add_argument('--user', type=str, required=True,
                            help='E-mail пользователя-создателя')
        parser.add_argument('--dry-run', action='store_true',
                            help='Только показать данные, не сохранять в БД')
        parser.add_argument('--force', action='store_true',
                            help='Удалить существующий проект с таким же названием и пересоздать')
        parser.add_argument('--sheet', type=str, default=None,
                            help='Название листа (необязательно, определяется автоматически)')

    def handle(self, *args, **options):
        from django.contrib.auth import get_user_model
        from budget.models import Project, BudgetSection, BudgetCategory

        User = get_user_model()
        filepath = options['filepath']
        dry_run = options['dry_run']
        force = options['force']

        try:
            user = User.objects.get(email=options['user'])
        except User.DoesNotExist:
            raise CommandError(f"Пользователь '{options['user']}' не найден в БД.")

        self.stdout.write(f"📂 Читаю файл: {filepath}")
        data = parse_excel(filepath, sheet_name=options.get('sheet'))

        self.stdout.write(self.style.SUCCESS('\n─── ПРОЕКТ ───'))
        self.stdout.write(f"  Название:  {data['name']}")
        self.stdout.write(f"  Донор:     {data['donor']}")
        self.stdout.write(f"  Код:       {data['project_code']}")
        self.stdout.write(f"  Период:    {data['start_date']} → {data['end_date']}")
        self.stdout.write(f"  Бюджет:    {data['total_budget']} USD")

        total_cats = 0
        for sec in data['sections']:
            cats = sec['categories']
            total_cats += len(cats)
            sec_total = sum(c['allocated_amount'] for c in cats)
            self.stdout.write(
                f"\n  📁 Секция {sec['code']}: {sec['name']} "
                f"({len(cats)} статей, ${sec_total:,.2f})"
            )
            for cat in cats:
                self.stdout.write(
                    f"     [{cat['code']:8s}] {cat['name'][:50]:<50}  ${cat['allocated_amount']:,.2f}"
                )

        self.stdout.write(f"\n  Итого статей: {total_cats}")

        if dry_run:
            self.stdout.write(self.style.WARNING('\n⚠️  --dry-run: в БД ничего не записано.'))
            return

        with transaction.atomic():
            existing = Project.objects.filter(
                name=data['name'],
                start_date=data['start_date']
            )
            if existing.exists():
                if force:
                    existing.delete()
                    self.stdout.write(self.style.WARNING('\n🗑️  Старый проект удалён.'))
                else:
                    raise CommandError(
                        "Проект уже существует. Добавь --force чтобы удалить его и пересоздать."
                    )

            project = Project.objects.create(
                name=data['name'],
                donor=data['donor'],
                project_code=data['project_code'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                total_budget=data['total_budget'],
                currency='USD',
                created_by=user,
            )
            self.stdout.write(f"\n✅ Создан проект #{project.pk}: {project.name}")

            for sec_order, sec in enumerate(data['sections']):
                section = BudgetSection.objects.create(
                    project=project,
                    code=sec['code'],
                    name=sec['name'],
                    order=sec_order,
                )
                for cat in sec['categories']:
                    BudgetCategory.objects.create(
                        project=project,
                        section=section,
                        code=cat['code'],
                        name=cat['name'],
                        unit_measure=cat['unit_measure'],
                        quantity=cat['quantity'],
                        unit_cost=cat['unit_cost'],
                        frequency=cat['frequency'],
                        allocated_amount=cat['allocated_amount'],
                        notes=cat['notes'],
                        order=cat['order'],
                    )
                self.stdout.write(
                    f"   ✓ Секция {sec['code']}: {sec['name']} ({len(sec['categories'])} статей)"
                )

        self.stdout.write(self.style.SUCCESS(
            f'\n🎉 Готово! Открой: http://localhost:8000/budget/{project.pk}/'
        ))
