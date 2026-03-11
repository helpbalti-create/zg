"""
Excel export — точное воспроизведение формата UMCOR/CWS:
  Лист 1: Detail Budget   (как REQ ZG UMCOR 4 _RU)
  Лист 2: Summary Report  (как XXXXX)
  Лист 3: Financial Report (как XXXXFinRep)
  Лист 4: Code Summary      (группировка по кодам статей)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django.db.models import Sum


# ── палитра (точно как в оригинале) ─────────────────────────────────────────
NAV  = 'FF1F3864'   # тёмно-синий заголовок
WHT  = 'FFFFFFFF'
SEC_BG   = 'FFBDD7EE'  # голубой — строка секции
SEC_FT   = 'FF1F3864'
SUB_BG   = 'FFDDEBF7'  # чуть светлее — подытог
WARN_BG  = 'FFFFF2CC'  # жёлтый 80–99%
OVER_BG  = 'FFFFC7CE'  # красный перерасход
OK_BG    = 'FFE2EFDA'  # зелёный норма
GREY     = 'FFF2F2F2'
LGREY    = 'FFF8F9FA'


def _f(bold=False, sz=10, color='FF000000', name='Arial'):
    return Font(bold=bold, size=sz, color=color, name=name)

def _fill(c):
    return PatternFill('solid', fgColor=c)

def _b(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

MONEY = '#,##0.00;(#,##0.00);"-"'
NUM2  = '0.00'


def _hdr(ws, r, c, v, bg=NAV, fg=WHT, bold=True, align='center', wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _f(bold=bold, color=fg)
    cell.fill = _fill(bg)
    cell.border = _b()
    cell.alignment = _al(align, 'center', wrap)
    return cell


def _cell(ws, r, c, v, bg=WHT, bold=False, align='left', fmt=None, color='FF000000'):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _f(bold=bold, color=color)
    cell.fill = _fill(bg)
    cell.border = _b()
    cell.alignment = _al(align, 'center', True)
    if fmt:
        cell.number_format = fmt
    return cell


def _money(ws, r, c, v, bg=WHT, bold=False, color='FF000000'):
    val = float(v) if v else 0
    cell = _cell(ws, r, c, val, bg=bg, bold=bold, align='right', fmt=MONEY, color=color)
    return cell


def export_project_to_excel(project, output):
    wb = Workbook()
    _sheet_detail(wb, project)
    _sheet_summary(wb, project)
    _sheet_finrep(wb, project)
    _sheet_code_summary(wb, project)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(output)


# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 1 — DETAIL BUDGET (как REQ ZG UMCOR 4 _RU)
# Колонки: # | CODE | BUDGET DESCRIPTION | UNIT MEASURE | QTY | UNIT COST | FREQ | TOTAL | BUDGET NOTES
# ════════════════════════════════════════════════════════════════════════════
def _sheet_detail(wb, project):
    ws = wb.create_sheet('Detail Budget')

    # Ширины колонок (A..I)
    widths = [6, 12, 45, 14, 10, 12, 10, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок проекта
    def title_row(r, text, size=11, bg=NAV, fg=WHT):
        ws.merge_cells(f'A{r}:I{r}')
        c = ws.cell(row=r, column=1, value=text)
        c.font = _f(bold=True, sz=size, color=fg)
        c.fill = _fill(bg)
        c.alignment = _al('center', 'center')
        ws.row_dimensions[r].height = 18 if size > 10 else 15

    title_row(1, 'Church World Service', 13)
    title_row(2, 'DETAIL BUDGET', 12)
    title_row(3, f'Project: {project.name}', 10, bg='FFDBE5F1', fg=SEC_FT)
    title_row(4, f'Project Period: {project.start_date.strftime("%d.%m.%Y")} — {project.end_date.strftime("%d.%m.%Y")}   |   Donor: {project.donor or "—"}   |   Code: {project.project_code or "—"}', 10, bg='FFDBE5F1', fg=SEC_FT)
    ws.merge_cells('A5:I5')  # пустая строка

    # Заголовки INCOME (копируем структуру оригинала)
    r = 6
    _hdr(ws, r, 1, 'CODE', align='center')
    ws.merge_cells(f'B{r}:H{r}')
    _hdr(ws, r, 2, 'DESCRIPTION')
    _hdr(ws, r, 9, 'APPROVED INCOME')

    income_rows = [
        ('UMCOR', 'UMCOR Sources', project.total_budget),
    ]
    for code, desc, amt in income_rows:
        r += 1
        _cell(ws, r, 1, code, bg=LGREY, bold=True, align='center')
        ws.merge_cells(f'B{r}:H{r}')
        _cell(ws, r, 2, desc, bg=LGREY)
        _money(ws, r, 9, amt, bg=LGREY, bold=True)

    r += 1
    ws.merge_cells(f'A{r}:H{r}')
    _cell(ws, r, 1, 'TOTAL INCOME', bg=SUB_BG, bold=True, align='right')
    _money(ws, r, 9, project.total_budget, bg=SUB_BG, bold=True)

    r += 2  # пустые строки

    # ── Заголовки основной таблицы ──────────────────────────────────────────
    r += 1
    headers = ['#', 'CODE', 'BUDGET DESCRIPTION', 'UNIT MEASURE', 'QUANTITY', 'UNIT COST', 'FREQ / MONTH', 'TOTAL', 'BUDGET NOTES']
    for ci, h in enumerate(headers, 1):
        _hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 28

    line = 1
    sections = project.sections.prefetch_related('categories__expenses').all()

    # Палитра цветов для групп кодов (чередуются)
    CODE_PALETTE = [
        'FFEAF4FB','FFFEF9E7','FFEAFAF1','FFFDEDEC','FFF4ECF7',
        'FFE8F8F5','FFFEF5E7','FFEBF5FB','FFF9EBEA','FFE9F7EF',
        'FFFDFEFE','FFF0F3FF','FFFFF8E7','FFF5EEF8',
    ]

    for section in sections:
        # Строка секции
        r += 1
        ws.merge_cells(f'A{r}:I{r}')
        sc = ws.cell(row=r, column=1, value=f'SECTION {section.code}:  {section.name.upper()}')
        sc.font = _f(bold=True, sz=10, color=SEC_FT)
        sc.fill = _fill(SEC_BG)
        sc.border = _b()
        sc.alignment = _al('left', 'center')
        ws.row_dimensions[r].height = 16

        # Группируем категории по коду
        from collections import OrderedDict
        code_map = OrderedDict()
        for cat in section.categories.all():
            code = (cat.code or '').strip() or None
            if code not in code_map:
                code_map[code] = []
            code_map[code].append(cat)

        pal_idx = 0
        for code, cats in code_map.items():
            grp_color = CODE_PALETTE[pal_idx % len(CODE_PALETTE)]
            pal_idx += 1
            multi = len(cats) > 1

            # Суммарный лимит/расход по коду
            grp_alloc = sum(c.allocated_amount for c in cats)
            grp_spent = sum(c.total_spent for c in cats)
            grp_rem   = grp_alloc - grp_spent
            grp_over  = grp_spent > grp_alloc
            grp_pct   = int(grp_spent / grp_alloc * 100) if grp_alloc else 0

            # Заголовок группы (только если несколько строк)
            if multi:
                r += 1
                ws.merge_cells(f'A{r}:F{r}')
                gh = ws.cell(row=r, column=1,
                    value=f'  ▸ Код {code or "—"}  ({len(cats)} строки)  |  Лимит: ${grp_alloc:,.2f}  |  Потрачено: ${grp_spent:,.2f}  |  Остаток: ${grp_rem:,.2f}  ({grp_pct}%)')
                gh.font = _f(bold=True, sz=9,
                    color=('FFCC0000' if grp_over else ('FF856404' if grp_pct >= 80 else '1F3864')))
                gh.fill = _fill(grp_color)
                gh.border = _b('thin')
                gh.alignment = _al('left', 'center')
                # Правые ячейки шапки группы
                for ci in range(7, 10):
                    ws.cell(row=r, column=ci).fill = _fill(grp_color)
                    ws.cell(row=r, column=ci).border = _b('thin')
                ws.row_dimensions[r].height = 13

            for cat in cats:
                r += 1
                spent = cat.total_spent
                if cat.is_over_budget:
                    bg = OVER_BG
                elif cat.is_warning:
                    bg = WARN_BG
                else:
                    bg = grp_color  # цвет группы

                _cell(ws, r, 1, line, bg=bg, align='center')
                _cell(ws, r, 2, cat.code or '—', bg=bg, align='center')
                _cell(ws, r, 3, cat.name, bg=bg)
                _cell(ws, r, 4, cat.unit_measure or '—', bg=bg, align='center')
                _money(ws, r, 5, cat.quantity, bg=bg)
                _money(ws, r, 6, cat.unit_cost, bg=bg)
                _money(ws, r, 7, cat.frequency, bg=bg)
                _money(ws, r, 8, cat.allocated_amount, bg=bg, bold=True)
                _cell(ws, r, 9, cat.notes or '', bg=bg)
                ws.row_dimensions[r].height = 14
                line += 1

            # Подытог группы (только если несколько строк)
            if multi:
                r += 1
                ws.merge_cells(f'A{r}:G{r}')
                gf = ws.cell(row=r, column=1, value=f'∑ Итого по коду {code}')
                gf.font = _f(bold=True, sz=9, color=SEC_FT)
                gf.fill = _fill(grp_color)
                gf.border = _b('thin')
                gf.alignment = _al('right', 'center')
                _money(ws, r, 8, grp_alloc, bg=grp_color, bold=True)
                _cell(ws, r, 9, f'{grp_pct}%', bg=grp_color, bold=True, align='center')
                ws.row_dimensions[r].height = 12

        # Подытог секции
        r += 1
        ws.merge_cells(f'A{r}:G{r}')
        stc = ws.cell(row=r, column=1, value=f'Sub-Total: Section {section.code}')
        stc.font = _f(bold=True, color=SEC_FT)
        stc.fill = _fill(SUB_BG)
        stc.border = _b()
        stc.alignment = _al('right', 'center')
        _money(ws, r, 8, section.total_allocated, bg=SUB_BG, bold=True)
        _cell(ws, r, 9, '', bg=SUB_BG)

    # Статьи без секции
    cats_no_sec = project.categories.filter(section__isnull=True)
    if cats_no_sec.exists():
        r += 1
        ws.merge_cells(f'A{r}:I{r}')
        ws.cell(row=r, column=1, value='OTHER ITEMS').font = _f(bold=True)
        for cat in cats_no_sec:
            r += 1
            _cell(ws, r, 1, line, align='center')
            _cell(ws, r, 2, cat.code or '—', align='center')
            _cell(ws, r, 3, cat.name)
            _cell(ws, r, 4, cat.unit_measure or '—', align='center')
            _money(ws, r, 5, cat.quantity)
            _money(ws, r, 6, cat.unit_cost)
            _money(ws, r, 7, cat.frequency)
            _money(ws, r, 8, cat.allocated_amount, bold=True)
            _cell(ws, r, 9, cat.notes or '')
            line += 1

    # TOTAL GRANT BUDGET
    r += 1
    ws.merge_cells(f'A{r}:G{r}')
    gt = ws.cell(row=r, column=1, value='TOTAL GRANT BUDGET  (A, B, C & D):')
    gt.font = _f(bold=True, sz=11, color=WHT)
    gt.fill = _fill(NAV)
    gt.border = _b()
    gt.alignment = _al('right', 'center')
    mc = _money(ws, r, 8, project.total_budget, bg=NAV, bold=True)
    mc.font = _f(bold=True, sz=11, color=WHT)
    _cell(ws, r, 9, '', bg=NAV)
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = 'A12'


# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 2 — SUMMARY REPORT (как XXXXX)
# Колонки: SECTION | CODE | DESCRIPTION | APPROVED BUDGET | TOTAL EXPENSES | BALANCE
# ════════════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, project):
    ws = wb.create_sheet('Summary Report')

    widths = [12, 10, 40, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, (text, bg, fg, sz) in enumerate([
        ('SUMMARY REPORT', NAV, WHT, 13),
        (f'{project.name}', 'FFDBE5F1', SEC_FT, 10),
        (f'{project.start_date.strftime("%d.%m.%Y")} — {project.end_date.strftime("%d.%m.%Y")}   |   {project.donor or ""}', 'FFDBE5F1', SEC_FT, 10),
    ], 1):
        ws.merge_cells(f'A{r}:F{r}')
        c = ws.cell(row=r, column=1, value=text)
        c.font = _f(bold=True, sz=sz, color=fg)
        c.fill = _fill(bg)
        c.alignment = _al('center', 'center')

    r = 5
    hdrs = ['SECTION', 'CODE', 'DESCRIPTION', f'APPROVED BUDGET ({project.currency})', f'TOTAL EXPENSES ({project.currency})', f'BALANCE ({project.currency})']
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 28

    for section in project.sections.prefetch_related('categories__expenses').all():
        r += 1
        ws.merge_cells(f'A{r}:F{r}')
        sc = ws.cell(row=r, column=1, value=f'SECTION {section.code}: {section.name.upper()}')
        sc.font = _f(bold=True, color=SEC_FT)
        sc.fill = _fill(SEC_BG)
        sc.border = _b()
        sc.alignment = _al('left', 'center')

        for cat in section.categories.all():
            r += 1
            spent = cat.total_spent
            rem = cat.remaining
            bg = OVER_BG if cat.is_over_budget else (WARN_BG if cat.is_warning else WHT)
            _cell(ws, r, 1, section.code, bg=bg, align='center')
            _cell(ws, r, 2, cat.code or '—', bg=bg, align='center')
            _cell(ws, r, 3, cat.name, bg=bg)
            _money(ws, r, 4, cat.allocated_amount, bg=bg)
            _money(ws, r, 5, spent, bg=bg)
            _money(ws, r, 6, rem, bg=(OVER_BG if rem < 0 else bg), bold=(rem < 0),
                   color=('FFCC0000' if rem < 0 else 'FF000000'))

        # Subtotal
        r += 1
        ws.merge_cells(f'A{r}:C{r}')
        st = ws.cell(row=r, column=1, value=f'TOTAL SECTION {section.code}')
        st.font = _f(bold=True, color=SEC_FT)
        st.fill = _fill(SUB_BG)
        st.border = _b()
        st.alignment = _al('right', 'center')
        s_rem = section.total_allocated - section.total_spent
        _money(ws, r, 4, section.total_allocated, bg=SUB_BG, bold=True)
        _money(ws, r, 5, section.total_spent, bg=SUB_BG, bold=True)
        _money(ws, r, 6, s_rem, bg=SUB_BG, bold=True)

    # Grand total
    r += 2
    ws.merge_cells(f'A{r}:C{r}')
    gt = ws.cell(row=r, column=1, value='TOTAL GRANT BUDGET (A, B, C & D)')
    gt.font = _f(bold=True, sz=11, color=WHT)
    gt.fill = _fill(NAV)
    gt.border = _b()
    gt.alignment = _al('right', 'center')
    total_spent = project.total_spent
    for ci, val in [(4, project.total_budget), (5, total_spent), (6, project.total_budget - total_spent)]:
        mc = _money(ws, r, ci, val, bg=NAV, bold=True)
        mc.font = _f(bold=True, sz=11, color=WHT)
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = 'A6'


# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 3 — FINANCIAL REPORT (как XXXXFinRep)
# Колонки: CODE | BUDGET DESCRIPTION | APPROVED BUDGET | EXP P1 | EXP P2 | EXP P3 | TOTAL EXP | BALANCE
# ════════════════════════════════════════════════════════════════════════════
def _sheet_finrep(wb, project):
    from .models import Expense
    ws = wb.create_sheet('Financial Report')

    widths = [12, 40, 18, 15, 15, 15, 18, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, (text, bg, fg, sz) in enumerate([
        ('FINANCIAL REPORT', NAV, WHT, 13),
        (f'{project.name}', 'FFDBE5F1', SEC_FT, 10),
        (f'Period: {project.start_date.strftime("%d.%m.%Y")} — {project.end_date.strftime("%d.%m.%Y")}   |   Sponsor: {project.donor or "—"}', 'FFDBE5F1', SEC_FT, 10),
    ], 1):
        ws.merge_cells(f'A{r}:H{r}')
        c = ws.cell(row=r, column=1, value=text)
        c.font = _f(bold=True, sz=sz, color=fg)
        c.fill = _fill(bg)
        c.alignment = _al('center', 'center')

    r = 5
    cur = project.currency
    hdrs = [
        'CODE', 'BUDGET DESCRIPTION',
        f'APPROVED BUDGET {cur}',
        f'EXPENSES PERIOD 1 {cur}',
        f'EXPENSES PERIOD 2 {cur}',
        f'EXPENSES PERIOD 3 {cur}',
        f'TOTAL EXPENSES {cur}',
        f'BALANCE {cur}',
    ]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 30

    def period_spent(cat, p):
        return Expense.objects.filter(category=cat, period=p).aggregate(
            t=Sum('amount'))['t'] or Decimal('0')

    for section in project.sections.prefetch_related('categories__expenses').all():
        r += 1
        ws.merge_cells(f'A{r}:H{r}')
        sc = ws.cell(row=r, column=1, value=f'SECTION {section.code}: {section.name.upper()}')
        sc.font = _f(bold=True, color=SEC_FT)
        sc.fill = _fill(SEC_BG)
        sc.border = _b()
        sc.alignment = _al('left', 'center')

        s_p1 = s_p2 = s_p3 = Decimal('0')
        for cat in section.categories.all():
            r += 1
            p1 = period_spent(cat, 1)
            p2 = period_spent(cat, 2)
            p3 = period_spent(cat, 3)
            total_exp = cat.total_spent
            rem = cat.remaining
            s_p1 += p1; s_p2 += p2; s_p3 += p3

            bg = OVER_BG if cat.is_over_budget else (WARN_BG if cat.is_warning else WHT)
            _cell(ws, r, 1, cat.code or '—', bg=bg, align='center')
            _cell(ws, r, 2, cat.name, bg=bg)
            _money(ws, r, 3, cat.allocated_amount, bg=bg)
            _money(ws, r, 4, p1 if p1 else None, bg=bg)
            _money(ws, r, 5, p2 if p2 else None, bg=bg)
            _money(ws, r, 6, p3 if p3 else None, bg=bg)
            _money(ws, r, 7, total_exp, bg=bg, bold=True)
            _money(ws, r, 8, rem, bg=(OVER_BG if rem < 0 else bg), bold=(rem < 0),
                   color=('FFCC0000' if rem < 0 else 'FF000000'))

        # Section subtotal
        r += 1
        ws.merge_cells(f'A{r}:B{r}')
        st = ws.cell(row=r, column=1, value=f'TOTAL SECTION {section.code}')
        st.font = _f(bold=True, color=SEC_FT)
        st.fill = _fill(SUB_BG)
        st.border = _b()
        st.alignment = _al('right', 'center')
        s_total = section.total_spent
        s_rem   = section.total_allocated - s_total
        for ci, val in [(3, section.total_allocated), (4, s_p1), (5, s_p2), (6, s_p3), (7, s_total), (8, s_rem)]:
            _money(ws, r, ci, val, bg=SUB_BG, bold=True)

    # Grand totals
    r += 2
    ws.merge_cells(f'A{r}:B{r}')
    gt = ws.cell(row=r, column=1, value='TOTAL GRANT BUDGET (A, B, C & D)')
    gt.font = _f(bold=True, sz=11, color=WHT)
    gt.fill = _fill(NAV)
    gt.border = _b()
    gt.alignment = _al('right', 'center')

    total_spent = project.total_spent
    all_p1 = Expense.objects.filter(category__project=project, period=1).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    all_p2 = Expense.objects.filter(category__project=project, period=2).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    all_p3 = Expense.objects.filter(category__project=project, period=3).aggregate(t=Sum('amount'))['t'] or Decimal('0')

    for ci, val in [
        (3, project.total_budget),
        (4, all_p1), (5, all_p2), (6, all_p3),
        (7, total_spent),
        (8, project.total_budget - total_spent),
    ]:
        mc = _money(ws, r, ci, val, bg=NAV, bold=True)
        mc.font = _f(bold=True, sz=11, color=WHT)
    ws.row_dimensions[r].height = 18

    # ── Детализация расходов ──────────────────────────────────────────────
    r += 3
    ws.merge_cells(f'A{r}:H{r}')
    dh = ws.cell(row=r, column=1, value='EXPENSE DETAILS / ДЕТАЛИЗАЦИЯ РАСХОДОВ')
    dh.font = _f(bold=True, sz=11, color=WHT)
    dh.fill = _fill(NAV)
    dh.border = _b()
    dh.alignment = _al('center', 'center')
    ws.row_dimensions[r].height = 18
    r += 1

    exp_hdrs = ['DATE', 'BUDGET LINE', 'DESCRIPTION', 'QTY', 'PERIOD', 'DOC #', f'AMOUNT {cur}', 'RECORDED AT', 'BY']
    for ci, h in enumerate(exp_hdrs, 1):
        _hdr(ws, r, ci, h)

    expenses = Expense.objects.filter(
        category__project=project
    ).select_related('category__section', 'created_by').order_by('date', 'created_at')

    for idx, exp in enumerate(expenses, 1):
        r += 1
        bg = GREY if idx % 2 == 0 else WHT
        _cell(ws, r, 1, exp.date.strftime('%d.%m.%Y'), bg=bg, align='center')
        _cell(ws, r, 2, f'{exp.category.code or "—"} — {exp.category.name[:40]}', bg=bg)
        _cell(ws, r, 3, exp.description, bg=bg)
        _money(ws, r, 4, getattr(exp, 'quantity', 1), bg=bg)
        _cell(ws, r, 5, f'Period {exp.period}', bg=bg, align='center')
        _cell(ws, r, 6, exp.document_number or '—', bg=bg, align='center')
        _money(ws, r, 7, exp.amount, bg=bg, bold=True)
        _cell(ws, r, 8, exp.created_at.strftime('%d.%m.%Y %H:%M'), bg=bg, align='center')
        _cell(ws, r, 9, str(exp.created_by) if exp.created_by else '—', bg=bg)

    # Итого расходов
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ft = ws.cell(row=r, column=1, value='TOTAL EXPENSES')
    ft.font = _f(bold=True, color=WHT)
    ft.fill = _fill(NAV)
    ft.border = _b()
    ft.alignment = _al('right', 'center')
    mc = _money(ws, r, 6, total_spent, bg=NAV, bold=True)
    mc.font = _f(bold=True, color=WHT)

    ws.freeze_panes = 'A6'


# ════════════════════════════════════════════════════════════════════════════
# ЛИСТ 4 — CODE SUMMARY (группировка статей по кодам)
# Показывает суммарный лимит и расходы на каждый код, даже если
# один код встречается в нескольких строках бюджета.
# ════════════════════════════════════════════════════════════════════════════
def _sheet_code_summary(wb, project):
    from collections import defaultdict
    from .models import Expense

    ws = wb.create_sheet('Code Summary')

    widths = [12, 42, 20, 20, 20, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок
    for r, (text, bg, fg, sz) in enumerate([
        ('CODE SUMMARY — УЧЁТ ПО КОДАМ', NAV, WHT, 13),
        (f'{project.name}', 'FFDBE5F1', SEC_FT, 10),
        (f'Период: {project.start_date.strftime("%d.%m.%Y")} — {project.end_date.strftime("%d.%m.%Y")}  |  Донор: {project.donor or "—"}', 'FFDBE5F1', SEC_FT, 10),
    ], 1):
        ws.merge_cells(f'A{r}:F{r}')
        c = ws.cell(row=r, column=1, value=text)
        c.font = _f(bold=True, sz=sz, color=fg)
        c.fill = _fill(bg)
        c.alignment = _al('center', 'center')

    r = 5
    cur = project.currency
    hdrs = [
        'КОД', 'СТАТЬИ БЮДЖЕТА (входят в код)',
        f'ЛИМИТ ({cur})', f'ПОТРАЧЕНО ({cur})',
        f'ОСТАТОК ({cur})', '% исп.'
    ]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 28

    # Группируем статьи по коду
    code_map = defaultdict(lambda: {'names': [], 'allocated': Decimal('0'), 'spent': Decimal('0')})
    for cat in project.categories.prefetch_related('expenses').all():
        code = (cat.code or '').strip()
        if not code:
            code = '(без кода)'
        code_map[code]['names'].append(cat.name)
        code_map[code]['allocated'] += cat.allocated_amount
        code_map[code]['spent'] += cat.total_spent

    prev_section_code = None
    for code in sorted(code_map.keys()):
        cs = code_map[code]
        allocated = cs['allocated']
        spent = cs['spent']
        remaining = allocated - spent
        pct = int(spent / allocated * 100) if allocated else 0

        is_over = spent > allocated
        is_warn = pct >= 80 and not is_over

        # Цвет строки
        if is_over:
            bg = OVER_BG
        elif is_warn:
            bg = WARN_BG
        elif pct >= 50:
            bg = OK_BG
        else:
            bg = WHT

        # Новая секция по первой букве кода (A, B, C, D, ADM, PRG...)
        section_prefix = code[:3] if len(code) >= 3 else code[:1]
        if section_prefix != prev_section_code:
            r += 1
            ws.merge_cells(f'A{r}:F{r}')
            sh = ws.cell(row=r, column=1, value=f'— {section_prefix}... —')
            sh.font = _f(bold=True, sz=9, color=SEC_FT)
            sh.fill = _fill(SEC_BG)
            sh.border = _b()
            sh.alignment = _al('center', 'center')
            ws.row_dimensions[r].height = 13
            prev_section_code = section_prefix

        r += 1
        names_str = ' / '.join(cs['names'])[:120]

        _cell(ws, r, 1, code, bg=bg, bold=True, align='center')
        _cell(ws, r, 2, names_str, bg=bg)
        _money(ws, r, 3, allocated, bg=bg)
        _money(ws, r, 4, spent, bg=bg, bold=is_over)
        rem_color = 'FFCC0000' if remaining < 0 else 'FF000000'
        _money(ws, r, 5, remaining, bg=bg, bold=is_over, color=rem_color)
        pct_cell = _cell(ws, r, 6, f'{pct}%', bg=bg, bold=is_over, align='center',
                         color=('FFCC0000' if is_over else ('FF856404' if is_warn else 'FF155724')))
        ws.row_dimensions[r].height = 14

    # Итоговая строка
    r += 2
    ws.merge_cells(f'A{r}:B{r}')
    gt = ws.cell(row=r, column=1, value='ИТОГО ПО ПРОЕКТУ')
    gt.font = _f(bold=True, sz=11, color=WHT)
    gt.fill = _fill(NAV)
    gt.border = _b()
    gt.alignment = _al('right', 'center')

    total_alloc = project.total_budget
    total_spent = project.total_spent
    total_rem = total_alloc - total_spent
    total_pct = int(total_spent / total_alloc * 100) if total_alloc else 0

    for ci, val in [(3, total_alloc), (4, total_spent), (5, total_rem)]:
        mc = _money(ws, r, ci, val, bg=NAV, bold=True)
        mc.font = _f(bold=True, sz=11, color=WHT)
    pct_c = _cell(ws, r, 6, f'{total_pct}%', bg=NAV, bold=True, align='center', color=WHT)
    ws.row_dimensions[r].height = 18

    # Легенда
    r += 2
    ws.merge_cells(f'A{r}:F{r}')
    leg = ws.cell(row=r, column=1, value='🔴 Перерасход (>100%)    🟡 Предупреждение (≥80%)    🟢 Норма (<80%)')
    leg.font = _f(sz=9, color='FF555555')
    leg.alignment = _al('center', 'center')

    ws.freeze_panes = 'A6'